"""Tests for the members-list export with selectable columns."""
import os
import sys
import tempfile

import pandas as pd

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from src.database import DatabaseManager
from src.reports import ReportGenerator


def _env():
    d = tempfile.mkdtemp()
    return DatabaseManager(os.path.join(d, "m.db")), d


def test_members_list_selected_columns_and_sorting():
    db, d = _env()
    db.add_individual("Bravo", "0712", "b@x", pf_no="PF-2", id_no="ID-2",
                      employment_status="Resigned")
    alpha = db.add_individual("Alpha", "0711", "a@x", pf_no="PF-1", id_no="ID-1")
    db.add_savings_transaction(alpha, "2026-01-01", "Deposit", 5000, "")

    out = os.path.join(d, "members.csv")
    ok, msg = ReportGenerator(db).generate_members_list(
        out, ["name", "pf_no", "employment_status", "savings", "is_retired"])
    assert ok, msg
    df = pd.read_csv(out)
    assert list(df.columns) == ["Name", "PF No", "Employment Status",
                                "Savings Balance", "Retired"]
    assert df.iloc[0]["Name"] == "Alpha"  # sorted by name
    assert df[df["Name"] == "Alpha"]["Savings Balance"].iloc[0] == 5000
    row_b = df[df["Name"] == "Bravo"].iloc[0]
    assert row_b["Employment Status"] == "Resigned"
    assert row_b["Retired"] == "Yes"      # non-Active status -> retired flag


def test_members_list_defaults_to_name_when_empty_selection():
    db, d = _env()
    db.add_individual("Solo", "0", "s@x")
    out = os.path.join(d, "m2.csv")
    ok, _ = ReportGenerator(db).generate_members_list(out, [])
    assert ok
    df = pd.read_csv(out)
    assert list(df.columns) == ["Name"]
    assert df.iloc[0]["Name"] == "Solo"


def test_members_list_respects_column_order():
    db, d = _env()
    db.add_individual("Jane", "0712", "j@x", pf_no="PF-1", id_no="ID-1")
    out = os.path.join(d, "ord.csv")
    ok, _ = ReportGenerator(db).generate_members_list(out, ["pf_no", "name", "id_no"])
    assert ok
    df = pd.read_csv(out)
    assert list(df.columns) == ["PF No", "Name", "ID No"]  # exact requested order


def test_members_list_excel_last_row_not_shaded():
    import openpyxl
    db, d = _env()
    db.add_individual("Alpha", "0", "a@x")
    db.add_individual("Bravo", "0", "b@x")
    out = os.path.join(d, "m.xlsx")
    ok, _ = ReportGenerator(db).generate_members_list(out, ["name"])
    assert ok
    ws = openpyxl.load_workbook(out).active
    last = ws.cell(row=ws.max_row, column=1)   # last member row
    assert "F0F0F0" not in str(last.fill.fgColor.rgb or "")  # no total shading
    assert last.font.bold is not True


def test_members_list_loan_balance_column():
    db, d = _env()
    from src.engine import LoanEngine
    ind = db.add_individual("Jane", "0", "j@x")
    LoanEngine(db).add_loan_event(ind, 100000, 12, "2025-01-01")
    out = os.path.join(d, "m3.csv")
    ok, _ = ReportGenerator(db).generate_members_list(out, ["name", "loan_balance"])
    assert ok
    df = pd.read_csv(out)
    assert df.iloc[0]["Loan Balance"] > 0  # outstanding principal+interest
