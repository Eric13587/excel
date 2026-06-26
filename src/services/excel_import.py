"""Excel (.xlsx) import: member rosters and fund-contribution sheets.

Two sheet shapes are supported, auto-detected from the header row:

* **roster**  – one member per row with any of Name / ID No / PF No / Phone /
  Email / Employment Status. Matched members are updated (fields backfilled);
  unmatched rows become new members.
* **contribution** – a Name + PF/Employee number + one column per month (e.g.
  the xmas sheet). Sets the member's PF number and imports the monthly amounts
  as Christmas or Benevolent deposits.

Names are reconciled with :class:`NameMatcher`. Nothing is written until
``apply`` is called; ``build_plan`` is side-effect free so the UI can preview.
"""
import re
from datetime import datetime

from src.services.name_matcher import NameMatcher
from src.services.christmas_service import ChristmasService
from src.services.benevolent_service import BenevolentService

_MONTH_NAMES = ['january', 'february', 'march', 'april', 'may', 'june',
                'july', 'august', 'september', 'october', 'november', 'december']


def _month_of(header):
    h = re.sub(r'[^a-z]', '', str(header).lower())[:3]
    for i, name in enumerate(_MONTH_NAMES, 1):
        if h and h == name[:3]:
            return i
    return None


def _norm_phone(v):
    return re.sub(r'\s+', '', str(v or '').strip())


def _clean(v):
    return str(v).strip() if v is not None else ''


class ExcelImporter:
    # logical field -> substrings that identify its header
    _FIELD_HINTS = {
        'name': ['name'],
        'id_no': ['idno', 'id no', 'id number', 'national id'],
        'pf_no': ['pf', 'employee number', 'employee no', 'emp no', 'payroll'],
        'phone': ['phone', 'mobile', 'tel'],
        'email': ['email', 'e-mail'],
        'employment_status': ['employment', 'status'],
    }

    def __init__(self, db):
        self.db = db

    # ------------------------------------------------------------------ #
    # Reading & detection
    # ------------------------------------------------------------------ #
    def read_sheet(self, path, sheet=None):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        data = list(ws.iter_rows(values_only=True))
        if not data:
            return [], []
        headers = [(_clean(h) or f"col{i}") for i, h in enumerate(data[0])]
        rows = [dict(zip(headers, r)) for r in data[1:]]
        return headers, rows

    def detect_mapping(self, headers):
        """Return {fields:{logical:header}, month_columns:[(header,month)], type}."""
        fields = {}
        used = set()
        lower = {h: h.lower() for h in headers}
        for field, hints in self._FIELD_HINTS.items():
            for h in headers:
                if h in used:
                    continue
                if any(hint in lower[h] for hint in hints):
                    fields[field] = h
                    used.add(h)
                    break
        month_columns = [(h, _month_of(h)) for h in headers if _month_of(h)]
        sheet_type = "contribution" if len(month_columns) >= 3 else "roster"
        return {"fields": fields, "month_columns": month_columns, "type": sheet_type}

    # ------------------------------------------------------------------ #
    # Planning (no writes)
    # ------------------------------------------------------------------ #
    def _candidates(self):
        return [(r[0], r[1]) for r in self.db.get_individuals()]

    def build_plan(self, rows, mapping, fund="christmas", fy_start_year=None):
        matcher = NameMatcher(self._candidates())
        fields = mapping["fields"]
        name_col = fields.get("name")
        plan_rows = []
        if not name_col:
            return {"type": mapping["type"], "rows": [], "error": "No Name column detected."}

        month_dates = None
        if mapping["type"] == "contribution":
            month_dates = self._month_date_map(mapping["month_columns"], fy_start_year)

        for r in rows:
            raw = _clean(r.get(name_col))
            if not raw or raw.upper() == "TOTAL":
                continue
            verdict, best = matcher.classify(raw)
            action = "update" if verdict in ("match", "review") else "create"
            entry = {"source_name": raw, "verdict": verdict, "match": best, "action": action}

            if mapping["type"] == "roster":
                entry["fields"] = self._row_fields(r, fields)
            else:
                entry["fields"] = {}
                if fields.get("pf_no"):
                    pf = _clean(r.get(fields["pf_no"]))
                    if pf:
                        entry["fields"]["pf_no"] = pf
                months = []
                for header, mnum in mapping["month_columns"]:
                    amt = self._num(r.get(header))
                    if amt > 0:
                        months.append((month_dates[mnum], amt))
                entry["months"] = months
                entry["fund"] = fund
            plan_rows.append(entry)
        return {"type": mapping["type"], "rows": plan_rows}

    def _row_fields(self, r, fields):
        out = {}
        for key in ("id_no", "pf_no", "employment_status"):
            if fields.get(key):
                v = _clean(r.get(fields[key]))
                if v:
                    out[key] = v
        if fields.get("phone"):
            ph = _norm_phone(r.get(fields["phone"]))
            if ph:
                out["phone"] = ph
        if fields.get("email"):
            em = _clean(r.get(fields["email"]))
            if em:
                out["email"] = em
        return out

    def _month_date_map(self, month_columns, fy_start_year):
        """Map each month number to a YYYY-MM-01 date. The first month column is
        the fiscal-year start; months before it roll into the next year."""
        if fy_start_year is None:
            fy_start_year = datetime.now().year
        start_month = month_columns[0][1] if month_columns else 1
        out = {}
        for _, m in month_columns:
            year = fy_start_year if m >= start_month else fy_start_year + 1
            out[m] = f"{year:04d}-{m:02d}-01"
        return out

    @staticmethod
    def _num(v):
        if v in (None, ""):
            return 0.0
        try:
            return float(str(v).replace(",", ""))
        except (TypeError, ValueError):
            return 0.0

    # ------------------------------------------------------------------ #
    # Apply (writes)
    # ------------------------------------------------------------------ #
    def apply(self, plan, fund="christmas", batch_id=None):
        """Execute a plan. Returns a stats/warnings summary."""
        if plan["type"] == "roster":
            return self._apply_roster(plan)
        return self._apply_contributions(plan, fund, batch_id)

    def _safe_unique(self, field, value, target_id, name, warnings):
        """Return value if free to assign, else '' (and warn). field: pf_no|id_no."""
        if not value:
            return value
        owner = (self.db.pf_no_owner if field == "pf_no" else self.db.id_no_owner)(value, exclude_id=target_id)
        if owner:
            warnings.append(f"{field.upper()} '{value}' for {name} already used by {owner} — skipped.")
            return ""
        return value

    def _apply_roster(self, plan):
        stats = {"updated": 0, "created": 0, "skipped": 0}
        warnings = []
        for e in plan["rows"]:
            if e["action"] == "skip":
                stats["skipped"] += 1
                continue
            f = dict(e["fields"])
            if e["action"] == "update" and e["match"]:
                tid = e["match"]["id"]
                f["pf_no"] = self._safe_unique("pf_no", f.get("pf_no", ""), tid, e["source_name"], warnings) if "pf_no" in f else None
                f["id_no"] = self._safe_unique("id_no", f.get("id_no", ""), tid, e["source_name"], warnings) if "id_no" in f else None
                cur = self.db.get_individual(tid) or {}
                self.db.update_individual(
                    tid, cur.get("name"), f.get("phone", cur.get("phone")),
                    f.get("email", cur.get("email")),
                    employment_status=f.get("employment_status"),
                    pf_no=f.get("pf_no"), id_no=f.get("id_no"))
                stats["updated"] += 1
            else:  # create
                pf = self._safe_unique("pf_no", f.get("pf_no", ""), None, e["source_name"], warnings)
                idn = self._safe_unique("id_no", f.get("id_no", ""), None, e["source_name"], warnings)
                self.db.add_individual(
                    e["source_name"], f.get("phone", ""), f.get("email", ""),
                    employment_status=f.get("employment_status", "Active") or "Active",
                    pf_no=pf, id_no=idn)
                stats["created"] += 1
        stats["warnings"] = warnings
        return stats

    def _apply_contributions(self, plan, fund, batch_id):
        stats = {"members": 0, "created": 0, "deposits": 0, "total": 0.0}
        warnings = []
        svc = ChristmasService(self.db) if fund == "christmas" else BenevolentService(self.db)
        for e in plan["rows"]:
            if e["action"] == "skip":
                continue
            if e["action"] == "update" and e["match"]:
                tid = e["match"]["id"]
            else:  # create the member from the sheet
                pf = self._safe_unique("pf_no", e["fields"].get("pf_no", ""), None, e["source_name"], warnings)
                tid = self.db.add_individual(e["source_name"], "", "", pf_no=pf)
                stats["created"] += 1
                e["fields"].pop("pf_no", None)  # already set on create
            # PF number for the matched/updated member (preserve phone/email)
            pf = e["fields"].get("pf_no")
            if pf:
                pf = self._safe_unique("pf_no", pf, tid, e["source_name"], warnings)
                if pf:
                    cur = self.db.get_individual(tid) or {}
                    self.db.update_individual(tid, cur.get("name"), cur.get("phone"),
                                              cur.get("email"), pf_no=pf)
            # contributions
            for date_str, amount in e.get("months", []):
                if fund == "christmas":
                    svc.add_deposit(tid, amount, date_str, notes="Excel import", batch_id=batch_id)
                else:
                    self.db.fund_add_transaction("benevolent_ledger", tid, date_str,
                                                 "Contribution", amount, "Excel import", batch_id)
                stats["deposits"] += 1
                stats["total"] += amount
            if e.get("months"):
                stats["members"] += 1
        stats["warnings"] = warnings
        return stats
